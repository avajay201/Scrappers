from bs4 import BeautifulSoup
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import uuid
from webdriver_manager.chrome import ChromeDriverManager


def get_html_selenium(url, scroll_pause=0.5, scroll_step=300):
    """
    Fetches the full HTML source of a webpage using headless Selenium with smooth scrolling.

    Args:
        url (str): The target URL.
        scroll_pause (float): Time to pause (in seconds) between scrolls.
        scroll_step (int): Number of pixels to scroll per step.
    """
    options = Options()
    options.add_argument("--headless")  # Run in background
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36")

    # Setup WebDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Open URL
    driver.get(url)
    
    # Wait for JavaScript content to load (Optional)
    driver.implicitly_wait(5)

    # Get total page height
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        for _ in range(0, last_height, scroll_step):
            driver.execute_script(f"window.scrollBy(0, {scroll_step});")
            time.sleep(scroll_pause)  # Pause for data to load

        # Allow time for new content to load
        time.sleep(1)

        # Calculate new scroll height after scrolling
        new_height = driver.execute_script("return document.body.scrollHeight")

        # Break if no new content is loaded
        if new_height == last_height:
            break
        last_height = new_height

    # Get full page source
    page_source = driver.page_source
    
    # Close driver
    driver.quit()
    
    return page_source

def scrape(url, products_scrape=10):
    """
    Scrapes product data from the given search URL and saves the page locally.

    Args:
        url (str): The search URL to scrape.
        products_scrape (int, optional): Number of products to scrape. Default is 10.

    Returns:
        bool: True if scraping is successful, False otherwise.
    """
    print(f"ğŸ” Scraping {products_scrape} {'products' if products_scrape > 1 else 'product'} from {url}...\n")

    try:
        # Generate a unique filename for storing the response
        file_name = f"{uuid.uuid4()}.html"
        file_path = os.path.join("temp_files", file_name)

        # Get html content to selenium
        html_content = get_html_selenium(url)
        with open(file_path, 'w') as f:
            f.write(html_content)

        # Call function to extract product details
        result = scrape_products(file_path, products_scrape)
        if not result:
            return result
        return True
    except Exception as err:
        print(f"âŒ [ERROR] Unexpected Error: {err}\n")
    return None

def scrape_products(file_path, products_scrape):
    """
    Extracts product information from the stored HTML file and saves it in Excel.

    Args:
        file_path (str): Path to the stored HTML file.
        products_scrape (int): Number of products to extract.
    """
    # Define HTML class names for product attributes
    products_container_class = "results-base"
    products_name_class = "product-product"
    products_price_class = "product-discountedPrice"
    products_rating_class = "product-ratingsContainer"

    # Read the saved HTML file
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    soup = BeautifulSoup(content, "html.parser")

    try:
        # Locate the product container
        products_container = soup.find(class_=products_container_class)

        products = products_container.children
        if not products:
            # Clean up the temporary file after scraping
            clean(file_path)
            return 0


        # Limit to requested count
        products = list(products)[:products_scrape]

        scrapped_products = []

        # Extract details for each product
        for product in products:
            product_name = product.find(class_=products_name_class).text if product.find(class_=products_name_class) else ""
            product_image = product.find("picture").find("img").attrs.get("src") if product.find("picture") and product.find("picture").find("img") else ""
            product_price = product.find(class_=products_price_class).text if product.find(class_=products_price_class) else ""
            product_rating = product.find(class_=products_rating_class).find("span").text if product.find(class_=products_rating_class) else ""
            product_link = product.find("a").attrs.get('href') if product.find("a") else ""

            if (product_name == "" and product_price == "" and product_link == "") or product_link == "":
                continue

            product_data = {
                "Name": product_name,
                "Image": product_image,
                "Price": product_price,
                "Rating": product_rating,
                "Link": url + '/' + product_link,
            }

            scrapped_products.append(product_data)

    except AttributeError as attr_err:
        print(f"âŒ Attribute Error: {attr_err}\n")
        if not scrapped_products:
            return None
    except Exception as err:
        print(f"âŒ Unexpected Error: {err}\n")
        return None

    # Clean up the temporary file after scraping
    clean(file_path)

    # Save extracted products to an Excel file
    save_to_excel(scrapped_products)
    print(f"âœ… {len(scrapped_products)} {'products' if products_scrape > 1 else 'product'} scrapped successfully.\n")
    return True

def save_to_excel(products):
    """
    ğŸ“Š Saves the scraped product data into an Excel file with improved formatting.

    Features:
    âœ… Auto-adjusts column widths based on content.
    âœ… Bold headers for better readability.
    âœ… Center-aligns the headers.

    Args:
        products (list of dict): List containing product details.
    """

    df = pd.DataFrame(products)

    # Generate a timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"scraped_products_{timestamp}.xlsx"

    # Save the DataFrame to Excel
    df.to_excel(file_name, index=False, engine="openpyxl")

    # Load the workbook to format it
    wb = load_workbook(file_name)
    ws = wb.active

    # Apply formatting
    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        # Determine max content width
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2  # Extra padding
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Limit max width

    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:  
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Wrap text in the "Description" column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Save the formatted file
    wb.save(file_name)
    print(f"ğŸ“ Data successfully saved and formatted in {file_name} âœ…\n")

def clean(file_path):
    """
    ğŸ—‘ï¸ Cleans up temporary files after scraping.

    Args:
        file_path (str): The path of the temporary file to delete.
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"ğŸ§¹ Successfully deleted temporary file: {file_path}\n")
        else:
            print(f"âš ï¸  File not found: {file_path}\n")
    except Exception as e:
        print(f"âŒ Error while deleting file {file_path}: {e}\n")


if __name__ == "__main__":
    # Ensure the 'temp_files' directory exists
    os.makedirs("temp_files", exist_ok=True)

    url = "https://www.myntra.com"

    # Get search query from user
    search_key = input("ğŸ” Enter a search key:\n")
    if not search_key:
        print("âŒ No search key provided!\n")
        quit()
    elif search_key.isdigit():
        print("âŒ Please enter a valid search term (string)!\n")
        quit()

    # Get the number of products to scrape (optional)
    products_scrape = input("ğŸ›ï¸  How many products do you want to scrape? (Optional):\n")
    if products_scrape:
        if not products_scrape.isdigit():
            print("âŒ Please enter a valid integer for product count!\n")
            quit()

        products_scrape = int(products_scrape)

        if products_scrape < 1:
            print("âŒ Please enter a number greater than 0!\n")
            quit()

        print(f"ğŸ”¢ You have chosen to scrape {products_scrape} products.\n")
    else:
        products_scrape = 10  # Default value if the user doesn't enter anything
        print("âš™ï¸  No input provided. Scraping default 10 products.\n")

    # Construct search URL
    search_url = f"{url}/{search_key}"

    # Call the scrape function with the provided number of products
    result = scrape(search_url, products_scrape)

    # Display final success/failure message
    if result:
        print("ğŸ‰ Scraping successful!\n")
    elif result == 0:
        print("âš ï¸  Products not found\n")
    else:
        print("âŒ Scraping failed.\n")
